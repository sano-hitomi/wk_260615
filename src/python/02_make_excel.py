import pandas as pd
import numpy as np
from scipy.stats import fisher_exact
from statsmodels.stats.multitest import multipletests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

# ─── データ読み込み ───
cand = pd.read_csv('/sessions/charming-youthful-pasteur/mnt/TMAO_pathway_analysis/data/plsda_mechanism_candidates.csv')
early = cand[cand['mechanism_class'] == 'Early_specific'].copy()
late  = cand[cand['mechanism_class'] == 'Late_specific'].copy()
rev   = cand[cand['mechanism_class'].isin(['Reversed_Q2','Reversed_Q4'])].copy()
background = cand.copy()
N = len(background)

# ─── Ontology ORA ───
def run_ora(subset, background, label):
    n = len(subset)
    N = len(background)
    results = []
    for cls, K in background['Ontology'].value_counts().items():
        if not cls or str(cls).strip() == '' or K < 2:
            continue
        k = (subset['Ontology'] == cls).sum()
        if k == 0:
            continue
        table = [[k, n-k],[K-k, N-n-(K-k)]]
        try:
            _, p = fisher_exact(table, alternative='greater')
        except:
            p = 1.0
        results.append({
            'Chemical_class': cls,
            'Count_in_set': int(k),
            'Set_size': int(n),
            'Count_in_bg': int(K),
            'BG_size': int(N),
            'Pct_in_set': round(100*k/n, 2),
            'Pct_in_bg': round(100*K/N, 2),
            'Fold_enrichment': round((k/n)/(K/N), 2) if K/N > 0 else 0,
            'p_value': p,
            'Group': label
        })
    df = pd.DataFrame(results)
    if len(df):
        _, fdr, _, _ = multipletests(df['p_value'], method='fdr_bh')
        df['FDR'] = fdr
    return df.sort_values('p_value').reset_index(drop=True)

ora_early = run_ora(early, background, 'Early_specific')
ora_late  = run_ora(late,  background, 'Late_specific')
ora_rev   = run_ora(rev,   background, 'Reversed')

# ─── Early vs Late 直接比較 ───
def ev_compare(early, late, min_count=3):
    all_cls = set(early['Ontology'].dropna().tolist()) | set(late['Ontology'].dropna().tolist())
    n_e, n_l = len(early), len(late)
    results = []
    for cls in sorted(all_cls):
        if not cls or str(cls).strip() == '':
            continue
        ke = (early['Ontology'] == cls).sum()
        kl = (late['Ontology']  == cls).sum()
        if ke + kl < min_count:
            continue
        table = [[ke, n_e-ke],[kl, n_l-kl]]
        try:
            _, p = fisher_exact(table)
        except:
            p = 1.0
        results.append({
            'Chemical_class': cls,
            'Early_count': int(ke), 'Early_n': int(n_e), 'Early_pct': round(100*ke/n_e, 2),
            'Late_count':  int(kl), 'Late_n':  int(n_l), 'Late_pct':  round(100*kl/n_l, 2),
            'Direction': 'Early_enriched' if ke/n_e > kl/n_l else 'Late_enriched',
            'Fold_ratio': round((ke/n_e)/(kl/n_l), 2) if kl/n_l > 0 else float('inf'),
            'p_value': p
        })
    df = pd.DataFrame(results)
    if len(df):
        _, fdr, _, _ = multipletests(df['p_value'], method='fdr_bh')
        df['FDR'] = fdr
    return df.sort_values('p_value').reset_index(drop=True)

ev_comp = ev_compare(early, late)

# ─── KEGG対応マップ ───
kegg_map = {
    'Dihydroxy bile acids, alcohols and derivatives': 'map00120 Primary bile acid biosynthesis; map04976 Bile secretion',
    'Trihydroxy bile acids, alcohols and derivatives': 'map00120 Primary bile acid biosynthesis',
    'Tetrahydroxy bile acids, alcohols and derivatives': 'map00120 Primary bile acid biosynthesis',
    'Monohydroxy bile acids, alcohols and derivatives': 'map00121 Secondary bile acid biosynthesis',
    'Glycinated bile acids and derivatives': 'map00120 Primary bile acid biosynthesis; map00121 Secondary bile acid biosynthesis',
    'Taurinated bile acids and derivatives': 'map00120 Primary bile acid biosynthesis; map00430 Taurine and hypotaurine metabolism',
    '7-alpha-hydroxysteroids': 'map00120 Primary bile acid biosynthesis; map00140 Steroid hormone biosynthesis',
    'Triacylglycerols': 'map00561 Glycerolipid metabolism; map04975 Fat digestion and absorption',
    'Phosphatidylcholines': 'map00564 Glycerophospholipid metabolism; map05231 Choline metabolism in cancer',
    'Phosphatidylethanolamines': 'map00564 Glycerophospholipid metabolism',
    '1-acyl-sn-glycero-3-phosphocholines': 'map00564 Glycerophospholipid metabolism',
    'Lysophosphatidylcholines': 'map00564 Glycerophospholipid metabolism',
    'Lysophosphatidylethanolamines': 'map00564 Glycerophospholipid metabolism',
    'Glycerophosphocholines': 'map00564 Glycerophospholipid metabolism',
    'N-acyl amines': 'map04723 Retrograde endocannabinoid signaling; map00590 Arachidonic acid metabolism',
    'N-acylethanolamines': 'map04723 Retrograde endocannabinoid signaling',
    'Glycosyl-N-acylsphingosines': 'map00600 Sphingolipid metabolism; map00601 Glycosphingolipid biosynthesis',
    'Ceramides': 'map00600 Sphingolipid metabolism; map04071 Sphingolipid signaling pathway',
    'Sphingomyelins': 'map00600 Sphingolipid metabolism',
    'Acyl CoAs': 'map00071 Fatty acid degradation; map01212 Fatty acid metabolism',
    'Very long-chain fatty acids': 'map01040 Biosynthesis of unsaturated fatty acids; map00071 Fatty acid degradation',
    'Long-chain fatty acids': 'map00071 Fatty acid degradation; map01212 Fatty acid metabolism',
    'Tetracarboxylic acids and derivatives': 'map00020 Citrate cycle (TCA cycle); map00630 Glyoxylate and dicarboxylate metabolism',
    'Triterpenoids': 'map00909 Sesquiterpenoid and triterpenoid biosynthesis; map00100 Steroid biosynthesis',
    'Steroidal saponins': 'map00100 Steroid biosynthesis',
    'Cholesterols and derivatives': 'map00100 Steroid biosynthesis; map04979 Cholesterol metabolism',
    'Xanthophylls': 'map00906 Carotenoid biosynthesis',
    'Proline and derivatives': 'map00330 Arginine and proline metabolism',
    'Glutamic acid and derivatives': 'map00250 Alanine, aspartate and glutamate metabolism',
    'Alpha amino acids': 'map00250 Alanine, aspartate and glutamate metabolism',
    'Flavin nucleotides': 'map00740 Riboflavin metabolism',
    'Eicosanoids': 'map00590 Arachidonic acid metabolism',
}

# ─── 代謝物リスト ───
def metabolite_list(df):
    cols = ['Alignment_ID','best_name','best_inchikey','Ontology','vip_max','loading_early','loading_late','mechanism_class']
    out = df[cols].copy()
    out['well_annotated'] = (~out['best_name'].str.startswith('low score', na=False)).astype(int)
    return out.sort_values(['well_annotated','vip_max'], ascending=[False,False]).reset_index(drop=True)

early_met = metabolite_list(early)
late_met  = metabolite_list(late)
rev_met   = metabolite_list(rev)

# ─── Excel作成 ───
wb = Workbook()

# スタイル定義
hdr_early = PatternFill('solid', start_color='D6E4F0')   # 水色
hdr_late  = PatternFill('solid', start_color='FDECEA')   # ピンク
hdr_rev   = PatternFill('solid', start_color='EAF4E0')   # 薄緑
hdr_comp  = PatternFill('solid', start_color='FFF3CD')   # 黄
hdr_summ  = PatternFill('solid', start_color='E8E4F8')   # 薄紫
bold = Font(bold=True, name='Arial', size=10)
normal = Font(name='Arial', size=10)
thin = Side(style='thin')
thin_border = Border(bottom=thin)

def style_header_row(ws, row, fill, font=None):
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font or bold
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

def set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

def df_to_ws(ws, df, fill, title=None, col_widths=None):
    if title:
        ws.append([title])
        ws[f'A{ws.max_row}'].font = Font(bold=True, name='Arial', size=12)
        ws.append([])
    ws.append(df.columns.tolist())
    style_header_row(ws, ws.max_row, fill)
    for row in df.itertuples(index=False):
        ws.append(list(row))
        for cell in ws[ws.max_row]:
            cell.font = normal
            cell.alignment = Alignment(wrap_text=False)
    if col_widths:
        set_col_widths(ws, col_widths)

# ─── Sheet 1: Summary ───
ws_sum = wb.active
ws_sum.title = 'Summary'
ws_sum['A1'] = 'TMAO Pathway Analysis — Early vs Late Summary'
ws_sum['A1'].font = Font(bold=True, name='Arial', size=14)
ws_sum['A2'] = f'Background: {N} features | Early_specific: {len(early)} | Late_specific: {len(late)} | Reversed: {len(rev)}'
ws_sum['A2'].font = Font(name='Arial', size=10, italic=True)

ws_sum.append([])
ws_sum['A4'] = '■ Early vs Late 有意差 Ontologyクラス (p < 0.05)'
ws_sum['A4'].font = Font(bold=True, name='Arial', size=11)

sig_comp = ev_comp[ev_comp['p_value'] < 0.05].copy()
ws_sum.append(['Chemical Class', 'Early Count', 'Early %', 'Late Count', 'Late %', 'Direction', 'p-value', 'FDR', 'KEGG Pathways'])
style_header_row(ws_sum, ws_sum.max_row, hdr_comp)
for _, r in sig_comp.iterrows():
    kegg = kegg_map.get(r['Chemical_class'], '—')
    ws_sum.append([r['Chemical_class'], r['Early_count'], r['Early_pct'],
                   r['Late_count'], r['Late_pct'], r['Direction'],
                   round(r['p_value'],4), round(r['FDR'],4), kegg])
    for cell in ws_sum[ws_sum.max_row]:
        cell.font = normal

ws_sum.append([])
ws_sum.append(['■ Early_specific トップ Ontologyクラス (p < 0.05)'])
ws_sum[f'A{ws_sum.max_row}'].font = Font(bold=True, name='Arial', size=11)
ws_sum.append(['Chemical Class', 'Count', 'Set %', 'BG %', 'Fold Enrich.', 'p-value', 'KEGG Pathways'])
style_header_row(ws_sum, ws_sum.max_row, hdr_early)
for _, r in ora_early[ora_early['p_value'] < 0.05].iterrows():
    kegg = kegg_map.get(r['Chemical_class'], '—')
    ws_sum.append([r['Chemical_class'], r['Count_in_set'], r['Pct_in_set'],
                   r['Pct_in_bg'], r['Fold_enrichment'], round(r['p_value'],4), kegg])
    for cell in ws_sum[ws_sum.max_row]: cell.font = normal

ws_sum.append([])
ws_sum.append(['■ Late_specific トップ Ontologyクラス (p < 0.05)'])
ws_sum[f'A{ws_sum.max_row}'].font = Font(bold=True, name='Arial', size=11)
ws_sum.append(['Chemical Class', 'Count', 'Set %', 'BG %', 'Fold Enrich.', 'p-value', 'KEGG Pathways'])
style_header_row(ws_sum, ws_sum.max_row, hdr_late)
for _, r in ora_late[ora_late['p_value'] < 0.05].iterrows():
    kegg = kegg_map.get(r['Chemical_class'], '—')
    ws_sum.append([r['Chemical_class'], r['Count_in_set'], r['Pct_in_set'],
                   r['Pct_in_bg'], r['Fold_enrichment'], round(r['p_value'],4), kegg])
    for cell in ws_sum[ws_sum.max_row]: cell.font = normal

set_col_widths(ws_sum, [45, 12, 10, 10, 10, 16, 10, 10, 70])

# ─── Sheet 2: Early vs Late 比較 ───
ws2 = wb.create_sheet('Early_vs_Late_Comparison')
ev_comp['KEGG_pathways'] = ev_comp['Chemical_class'].map(kegg_map).fillna('—')
df_to_ws(ws2, ev_comp, hdr_comp,
         title='Early vs Late — Chemical Class Enrichment Comparison (Fisher\'s exact test)',
         col_widths=[45, 12, 8, 10, 12, 8, 10, 16, 10, 10, 70])

# ─── Sheet 3: Early ORA ───
ws3 = wb.create_sheet('Early_specific_ORA')
ora_early['KEGG_pathways'] = ora_early['Chemical_class'].map(kegg_map).fillna('—')
df_to_ws(ws3, ora_early, hdr_early,
         title='Early_specific — Chemical Class ORA (enriched vs background)',
         col_widths=[45, 10, 8, 10, 8, 10, 10, 10, 10, 70])

# ─── Sheet 4: Late ORA ───
ws4 = wb.create_sheet('Late_specific_ORA')
ora_late['KEGG_pathways'] = ora_late['Chemical_class'].map(kegg_map).fillna('—')
df_to_ws(ws4, ora_late, hdr_late,
         title='Late_specific — Chemical Class ORA (enriched vs background)',
         col_widths=[45, 10, 8, 10, 8, 10, 10, 10, 10, 70])

# ─── Sheet 5: Reversed ORA ───
ws5 = wb.create_sheet('Reversed_ORA')
ora_rev['KEGG_pathways'] = ora_rev['Chemical_class'].map(kegg_map).fillna('—')
df_to_ws(ws5, ora_rev, hdr_rev,
         title='Reversed (Q2+Q4) — Chemical Class ORA (Early↔Late逆転候補)',
         col_widths=[45, 10, 8, 10, 8, 10, 10, 10, 10, 70])

# ─── Sheet 6: Early metabolites ───
ws6 = wb.create_sheet('Early_specific_Metabolites')
df_to_ws(ws6, early_met.head(200), hdr_early,
         title='Early_specific Metabolites (top 200, VIP順)',
         col_widths=[12, 55, 30, 35, 10, 12, 12, 18, 14])

# ─── Sheet 7: Late metabolites ───
ws7 = wb.create_sheet('Late_specific_Metabolites')
df_to_ws(ws7, late_met.head(200), hdr_late,
         title='Late_specific Metabolites (top 200, VIP順)',
         col_widths=[12, 55, 30, 35, 10, 12, 12, 18, 14])

# ─── Sheet 8: Reversed metabolites ───
ws8 = wb.create_sheet('Reversed_Metabolites')
rev_met2 = rev[['Alignment_ID','best_name','best_inchikey','Ontology',
                'vip_max','loading_early','loading_late','mechanism_class','mechanism_score']].copy()
rev_met2['well_annotated'] = (~rev_met2['best_name'].str.startswith('low score', na=False)).astype(int)
rev_met2 = rev_met2.sort_values(['well_annotated','vip_max'], ascending=[False,False]).reset_index(drop=True)
df_to_ws(ws8, rev_met2.head(100), hdr_rev,
         title='Reversed (Q2+Q4) Metabolites — EarlyとLateで逆方向 (機構差候補)',
         col_widths=[12, 55, 30, 35, 10, 12, 12, 18, 14, 14])

outpath = '/sessions/charming-youthful-pasteur/mnt/TMAO_pathway_analysis/output/TMAO_pathway_results.xlsx'
wb.save(outpath)
print(f"Saved: {outpath}")
